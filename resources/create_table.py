from openpyxl import load_workbook


wb = load_workbook(filename = 'Schemas.xlsx')


sheet = wb['Schema']
#print(sheet_ranges['A1'].value)
#for row in range(1, max_row + 1):
l_schema = {}
for row in sheet.iter_rows():
    l_sql = ""
    l_sql = "CREATE TABLE "
    if l_schema.get(row[0].value, None) == None:
        l_schema[row[0].value] = []
    #[寫法一]
    #l_dict = {}
    #l_dict['Col_Name'] =  row[1].value
    #l_dict['Type'] = row[2].value
    #l_dict['Primary_Key'] = row[3].value
    #l_dict['Description'] = row[4].value
    #l_schema[row[0].value].append(l_dict)
    #[寫法二]
    l_list = ['Col_Name', 'Type', 'Primary_Key', 'Description']
    l_dict = {i:row[idx+1].value for idx, i in enumerate(l_list)}
    l_schema[row[0].value].append(l_dict)
print(l_schema)
wb.close()
